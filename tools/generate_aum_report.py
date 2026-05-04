"""Generate AUM report by fetching live NAVs from AMFI and computing current value."""
import glob
import os
import sys

# Bootstrap Django so we can reuse the production NAV matcher.
_DJANGO_DIR = os.path.abspath(
    os.path.join(os.path.dirname(__file__), "..", "examples", "django_reference")
)
sys.path.insert(0, _DJANGO_DIR)
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "reference_app.settings")
os.environ.setdefault(
    "OFR_DATABASE_URL", "postgresql+psycopg://ofr:ofr@localhost:5438/ofr"
)
import django

django.setup()

import pandas as pd
import psycopg
import requests
from dbfread import DBF
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

# Use the production-grade fuzzy matcher that handles plan/option preferences
# and "Formerly Known As" rename annotations.
from clients.amfi_nav import lookup_nav as _lookup_nav  # noqa: E402

# === 1. Fetch AMFI NAV feed ===
print("Fetching AMFI NAV feed...")
resp = requests.get("https://www.amfiindia.com/spages/NAVAll.txt", timeout=30)
lines = resp.text.strip().split("\n")

amfi_rows = []
for line in lines:
    parts = line.strip().split(";")
    if len(parts) == 6:
        amfi_rows.append({
            "amfi_code": parts[0].strip(),
            "isin_growth": parts[1].strip(),
            "isin_reinvest": parts[2].strip(),
            "amfi_scheme_name": parts[3].strip(),
            "nav": parts[4].strip(),
            "nav_date": parts[5].strip(),
        })
amfi = pd.DataFrame(amfi_rows)
amfi["nav"] = pd.to_numeric(amfi["nav"], errors="coerce")
print(f"AMFI feed: {len(amfi)} schemes loaded")

# Build ISIN -> NAV lookup
isin_to_nav = {}
isin_to_name = {}
isin_to_date = {}
for _, row in amfi.iterrows():
    for isin_col in ["isin_growth", "isin_reinvest"]:
        isin = row[isin_col]
        if isin and isin != "-" and pd.notna(row["nav"]):
            isin_to_nav[isin] = row["nav"]
            isin_to_name[isin] = row["amfi_scheme_name"]
            isin_to_date[isin] = row["nav_date"]

# Build name lookup for fallback matching
amfi_name_lower = {}
for _, r in amfi.iterrows():
    key = r["amfi_scheme_name"].lower().strip()
    if pd.notna(r["nav"]):
        amfi_name_lower[key] = r


def find_nav_by_name(scheme_name):
    """Try to match scheme name against AMFI feed."""
    sn = scheme_name.lower().strip()
    if sn in amfi_name_lower:
        r = amfi_name_lower[sn]
        return r["nav"], r["nav_date"], r["amfi_scheme_name"]
    # Substring match
    matches = [(k, v) for k, v in amfi_name_lower.items() if sn[:30] in k]
    if len(matches) == 1:
        r = matches[0][1]
        return float(r["nav"]) if pd.notna(r["nav"]) else None, r["nav_date"], r["amfi_scheme_name"]
    return None, None, None


# === 2. Collect ISINs + scheme names from source files ===
scheme_isins = {}
scheme_names = {}

for f in glob.glob("examples/django_reference/uploaded_files/*.csv"):
    df = pd.read_csv(f, dtype=str)
    df.columns = [c.strip("'\" ") for c in df.columns]
    for col in df.columns:
        if df[col].dtype == object:
            df[col] = df[col].str.strip("'\" ")
    if "Scheme Code" in df.columns and "ISIN" in df.columns:
        for _, row in df[["Scheme Code", "ISIN"]].drop_duplicates().iterrows():
            sc = str(row["Scheme Code"]).strip()
            isin = str(row["ISIN"]).strip()
            if sc and isin and isin != "nan":
                scheme_isins[sc] = isin
    if "Scheme Code" in df.columns and "Fund Description" in df.columns:
        for _, row in df[["Scheme Code", "Fund Description"]].drop_duplicates().iterrows():
            sc = str(row["Scheme Code"]).strip()
            nm = str(row["Fund Description"]).strip()
            if sc and nm and nm != "nan":
                scheme_names[sc] = nm
    if "PRODCODE" in df.columns and "SCHEME" in df.columns:
        for _, row in df[["PRODCODE", "SCHEME"]].drop_duplicates().iterrows():
            sc = str(row["PRODCODE"]).strip()
            nm = str(row["SCHEME"]).strip()
            if sc and nm and nm != "nan":
                scheme_names[sc] = nm

for f in glob.glob("examples/django_reference/uploaded_files/*.dbf"):
    for row in DBF(f, load=True, char_decode_errors="ignore"):
        c = str(row.get("PRODCODE", "")).strip()
        n = str(row.get("SCHEME", "")).strip()
        if c and n:
            scheme_names[c] = n

# === 3. Query holdings ===
conn = psycopg.connect("postgresql://ofr:ofr@localhost:5438/ofr")
holdings = pd.read_sql(
    """
    SELECT
        a.name AS "Investor Name",
        a.pan AS "PAN",
        s.scheme_code AS "Scheme Code",
        COUNT(*) AS "Total Transactions",
        SUM(CASE WHEN t.action = 'buy' THEN t.units ELSE 0 END) -
        SUM(CASE WHEN t.action = 'sell' THEN t.units ELSE 0 END) AS "Net Units",
        SUM(CASE WHEN t.action = 'buy' THEN t.amount ELSE -t.amount END) AS "Net Invested (INR)",
        MIN(t.transaction_date) AS "First Transaction",
        MAX(t.transaction_date) AS "Last Transaction",
        STRING_AGG(DISTINCT sf.filename, ', ') AS "Source Files"
    FROM openreversefeed.transactions t
    JOIN openreversefeed.accounts a ON a.id = t.account_id
    JOIN openreversefeed.schemes s ON s.id = t.scheme_id
    JOIN openreversefeed.source_files sf ON sf.id = t.source_file_id
    WHERE t.transaction_date <= CURRENT_DATE
    GROUP BY a.name, a.pan, s.scheme_code
    ORDER BY a.name, s.scheme_code
    """,
    conn,
)
conn.close()

# === 4. Map scheme names, ISINs, NAVs ===
holdings.insert(
    3,
    "Scheme Name",
    holdings["Scheme Code"].map(scheme_names).fillna(holdings["Scheme Code"]),
)

nav_values = []
nav_dates = []
amfi_matched = []
for _, row in holdings.iterrows():
    sc = row["Scheme Code"]
    name = scheme_names.get(sc, sc)
    isin = scheme_isins.get(sc)
    # Production matcher: ISIN → exact → punct-normalized → fuzzy (plan-aware)
    nav_val, nav_dt, matched_name = _lookup_nav(sc, name, isin)
    nav_values.append(nav_val)
    nav_dates.append(nav_dt)
    amfi_matched.append(matched_name)

holdings["Net Units"] = holdings["Net Units"].astype(float).round(3)
holdings["Net Invested (INR)"] = holdings["Net Invested (INR)"].astype(float).round(2)
holdings["Current NAV"] = nav_values
holdings["NAV Date"] = nav_dates
holdings["AMFI Scheme Match"] = amfi_matched
holdings["Current Value (INR)"] = holdings.apply(
    lambda r: round(r["Net Units"] * r["Current NAV"], 2)
    if pd.notna(r["Current NAV"]) and r["Net Units"] > 0
    else None,
    axis=1,
)
holdings["P&L (INR)"] = holdings.apply(
    lambda r: round(r["Current Value (INR)"] - r["Net Invested (INR)"], 2)
    if pd.notna(r["Current Value (INR)"])
    else None,
    axis=1,
)

# === 5. Write Excel ===
import datetime as _dt
_today = _dt.date.today().strftime("%d%b%Y")
output = f"C:/Users/kavya/OneDrive/Desktop/AUM_Report_{_today}.xlsx"
with pd.ExcelWriter(output, engine="openpyxl") as writer:
    holdings.to_excel(writer, sheet_name="Holdings & AUM", index=False)
    ws = writer.sheets["Holdings & AUM"]

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for col_idx, col_cells in enumerate(
        ws.iter_cols(min_row=1, max_row=ws.max_row), 1
    ):
        max_len = 0
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 55)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            if isinstance(cell.value, float):
                header = str(ws.cell(1, cell.column).value)
                if "Units" in header or "NAV" in header:
                    cell.number_format = "#,##0.000"
                else:
                    cell.number_format = "#,##0.00"

    # Summary sheet
    matched_holdings = holdings[holdings["Current Value (INR)"].notna()].copy()
    total_invested = matched_holdings["Net Invested (INR)"].sum()
    total_value = matched_holdings["Current Value (INR)"].sum()
    total_pl = matched_holdings["P&L (INR)"].sum()
    unmatched = holdings[holdings["Current NAV"].isna()]["Scheme Code"].unique()

    pnl_pct = f"{round(total_pl / total_invested * 100, 2)}%" if total_invested else "N/A"

    summary_data = pd.DataFrame(
        [
            {"Metric": "Total Investors", "Value": holdings["PAN"].nunique()},
            {"Metric": "Total Schemes", "Value": len(holdings["Scheme Code"].unique())},
            {
                "Metric": "Schemes with NAV matched",
                "Value": int(holdings["Current NAV"].notna().sum()),
            },
            {"Metric": "Schemes without NAV", "Value": len(unmatched)},
            {"Metric": "Total Net Invested (INR)", "Value": round(total_invested, 2)},
            {
                "Metric": "Total Current Value / AUM (INR)",
                "Value": round(total_value, 2),
            },
            {"Metric": "Total P&L (INR)", "Value": round(total_pl, 2)},
            {"Metric": "P&L %", "Value": pnl_pct},
        ]
    )
    summary_data.to_excel(writer, sheet_name="Summary", index=False)

    if len(unmatched) > 0:
        unmatched_df = pd.DataFrame(
            {
                "Scheme Code": unmatched,
                "Scheme Name": [scheme_names.get(s, s) for s in unmatched],
            }
        )
        unmatched_df.to_excel(writer, sheet_name="Unmatched Schemes", index=False)

matched_count = holdings["Current NAV"].notna().sum()
print(f"Written: {output}")
print(f"Rows: {len(holdings)} | NAV matched: {matched_count}/{len(holdings)}")
print(f"Unmatched schemes: {list(unmatched)}")
print(f"Total AUM (matched only): INR {total_value:,.2f}")
print(f"Total P&L: INR {total_pl:,.2f} ({pnl_pct})")
